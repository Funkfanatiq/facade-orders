"""
Генератор ТОРГ-12 через openpyxl + Excel-шаблон.
Заполняет шаблон ТОРГ-12 (образец 11 от 12.07.2016) данными счёта.
Вывод: Excel (xlsx).
"""
import io
from pathlib import Path

# Путь к шаблону (рядом с этим модулем)
_TEMPLATE_PATH = Path(__file__).parent / "torg12_template.xlsx"

# Стиль рамок
_THIN = "thin"


def _apply_border(ws, min_row, min_col, max_row, max_col, style=_THIN):
    """Применить рамки ко всем ячейкам в диапазоне. MergedCell пропускаем (read-only)."""
    try:
        from openpyxl.styles import Border, Side
        from openpyxl.cell.cell import MergedCell
        side = Side(border_style=style, color="000000")
        border = Border(left=side, top=side, right=side, bottom=side)
        for row in ws.iter_rows(min_row=min_row, min_col=min_col, max_row=max_row, max_col=max_col):
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                cell.border = border
    except Exception:
        pass


def _fmt_num(x):
    try:
        return f"{float(x or 0):.2f}".replace(".", ",")
    except (TypeError, ValueError):
        return "0,00"


def _amount_to_words_rub(amount):
    """Сумма прописью: 92701.00 -> Девяносто две тысячи семьсот один рубль 00 копеек."""
    try:
        from num2words import num2words
        s = num2words(amount, lang="ru", to="currency", currency="RUB")
        return s.replace("-", " ").capitalize() if s else f"{amount:.2f}".replace(".", ",")
    except Exception:
        rub = int(amount)
        kop = round((amount - rub) * 100)
        return f"{rub} руб. {kop:02d} коп."


def _cell_top_left(ws, row, col):
    """Для merged cell возвращает (min_row, min_col); иначе (row, col)."""
    for m in ws.merged_cells.ranges:
        if m.min_row <= row <= m.max_row and m.min_col <= col <= m.max_col:
            return (m.min_row, m.min_col)
    return (row, col)


def _put_cell(ws, row, col, value):
    """Пишет в top-left merge, иначе в ячейку (обход MergedCell read-only)."""
    r, c = _cell_top_left(ws, row, col)
    ws.cell(row=r, column=c, value=value)


def _put_qty_price_sum(ws, row, col_qty, col_prc, col_sum, qty, prc, s, col_vat_rate=None, col_vat_amt=None, col_sum_vat=None):
    """Записать кол-во, цену, сумму, НДС. prc=None для строки «Всего»."""
    _put_cell(ws, row, col_qty, _fmt_num(qty))
    _put_cell(ws, row, col_prc, "х" if prc is None else _fmt_num(prc))
    _put_cell(ws, row, col_sum, _fmt_num(s))
    if col_vat_rate is not None:
        _put_cell(ws, row, col_vat_rate, "х" if prc is None else "0%")
    if col_vat_amt is not None:
        _put_cell(ws, row, col_vat_amt, _fmt_num(0))
    if col_sum_vat is not None:
        _put_cell(ws, row, col_sum_vat, _fmt_num(s))


def _org_string(config):
    """Полная строка реквизитов организации для грузоотправителя/поставщика."""
    parts = [
        config.get("COMPANY_NAME", ""),
        config.get("COMPANY_ADDRESS", ""),
    ]
    inn = config.get("COMPANY_INN", "")
    kpp = config.get("COMPANY_KPP", "")
    if inn:
        parts.append(f"ИНН {inn}")
    if kpp:
        parts.append(f"КПП {kpp}")
    acc = config.get("COMPANY_ACCOUNT", "")
    bank = config.get("COMPANY_BANK", "")
    if acc and bank:
        parts.append(f"р/с {acc} в {bank}")
    bik = config.get("COMPANY_BIK", "")
    corr = config.get("COMPANY_CORR_ACCOUNT", "")
    if bik:
        parts.append(f"БИК {bik}")
    if corr:
        parts.append(f"к/с {corr}")
    return ", ".join(p for p in parts if p)


def _buyer_string(counterparty):
    """Реквизиты грузополучателя/плательщика."""
    parts = [counterparty.full_name or counterparty.name]
    addr = counterparty.address or counterparty.legal_address
    if addr:
        parts.append(addr)
    if counterparty.inn:
        parts.append(f"ИНН {counterparty.inn}")
    if counterparty.kpp:
        parts.append(f"КПП {counterparty.kpp}")
    if counterparty.payment_account or counterparty.bank:
        acc = counterparty.payment_account or ""
        bank = counterparty.bank or ""
        if acc and bank:
            parts.append(f"р/с {acc} в {bank}")
        elif acc:
            parts.append(f"р/с {acc}")
        elif bank:
            parts.append(f"в {bank}")
    if counterparty.bik:
        parts.append(f"БИК {counterparty.bik}")
    if counterparty.corr_account:
        parts.append(f"к/с {counterparty.corr_account}")
    return ", ".join(p for p in parts if p)


def generate_torg12_xlsx(invoice, counterparty, config, template_path=None):
    """
    Заполняет Excel-шаблон ТОРГ-12 и возвращает BytesIO с xlsx.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError("Установите openpyxl: pip install openpyxl")

    path = Path(template_path or _TEMPLATE_PATH)
    if not path.exists():
        raise FileNotFoundError(f"Шаблон ТОРГ-12 не найден: {path}")

    wb = load_workbook(path)
    ws = wb.active

    inv_num = str(invoice.invoice_number or "")
    inv_dt = (invoice.invoice_date or __import__("datetime").date.today()).strftime("%d.%m.%Y")
    # Дата составления = дата создания документа ТОРГ-12 (сегодня)
    doc_dt_str = __import__("datetime").date.today().strftime("%d.%m.%Y")
    basis = f"Счет на оплату № {inv_num} от {inv_dt}"

    org = _org_string(config)
    buyer = _buyer_string(counterparty)

    # Верхний блок — _put_cell для обхода MergedCell
    _put_cell(ws, 3, 2, org)   # B3
    _put_cell(ws, 6, 2, buyer)  # B6
    _put_cell(ws, 10, 4, org)   # D10
    _put_cell(ws, 12, 4, buyer)  # D12
    _put_cell(ws, 14, 4, basis)  # D14

    # Номер документа = номер счёта, Дата составления = день отгрузки
    _put_cell(ws, 17, 11, inv_num)   # K17
    _put_cell(ws, 17, 15, doc_dt_str)  # O17

    # Правый блок «Коды»: AM = колонка 39
    okpo = config.get("COMPANY_OKPO") or ""
    cp_okpo = getattr(counterparty, "okpo", None) or ""
    _put_cell(ws, 3, 39, "0330212")
    if okpo:
        _put_cell(ws, 4, 39, okpo)
    if cp_okpo:
        _put_cell(ws, 7, 39, cp_okpo)
    if okpo:
        _put_cell(ws, 9, 39, okpo)
    if cp_okpo:
        _put_cell(ws, 11, 39, cp_okpo)
    _put_cell(ws, 13, 39, inv_num)

    # Таблица — по указанию пользователя: 22=Количество, 24=Цена, 26=Общая сумма; код по ОКЕИ 055
    COL_N = 2       # B — Номер по порядку
    COL_NAME = 3    # C — наименование, характеристика, сорт, артикул товара
    COL_UNIT = 8    # H — Единица измерения, наименование (кв.м)
    # COL_OKEI = 10 — J в merge H23:K23, писать нельзя; пишем "055" в единицу (топ-левая H)
    # Merge row 23: V-W(22), X-Y(24), Z-AD(26), AE-AH(31), AI-AK(35), AL-AO(38)
    COL_QTY = 22    # V — Количество
    COL_PRC = 24    # X — Цена
    COL_SUM = 26    # Z — Общая сумма
    COL_VAT_RATE = 31   # AE — ставка НДС (28 в merge Z-AD → пишем в top-left 31)
    COL_VAT_AMT = 35    # AI — сумма НДС (33 в merge AE-AH → top-left 35)
    COL_SUM_VAT = 38    # AL — Сумма с НДС
    DATA_START_ROW = 23
    DEFAULT_DATA_ROWS = 6
    total_sum = 0.0
    total_qty = 0.0

    items = list(invoice.items)
    insert_count = max(0, len(items) - DEFAULT_DATA_ROWS)
    if insert_count > 0:
        insert_at = DATA_START_ROW + DEFAULT_DATA_ROWS
        ws.insert_rows(insert_at, insert_count)
        for k in range(insert_count):
            r = insert_at + k
            for start, end in [(3, 6), (8, 11), (18, 21)]:  # без merge N-P: O,P — отдельные ячейки
                try:
                    ws.merge_cells(start_row=r, start_column=start, end_row=r, end_column=end)
                except Exception:
                    pass

    for i, it in enumerate(items):
        row = DATA_START_ROW + i
        qty = float(it.quantity or 0)
        prc = float(it.price or 0)
        s = round(qty * prc, 2)
        total_sum += s
        total_qty += qty

        _put_cell(ws, row, COL_N, i + 1)
        _put_cell(ws, row, COL_NAME, str(it.name or ""))
        # Ед.изм: H — топ merge H23:K23; код ОКЕИ 055 — J в merge, пишем в H
        unit = str(it.unit or "кв.м")
        sqm = ("кв" in unit.lower() or "м²" in unit or unit == "м2")
        _put_cell(ws, row, COL_UNIT, f"{unit} (055)" if sqm else unit)
        _put_qty_price_sum(ws, row, COL_QTY, COL_PRC, COL_SUM, qty, prc, s,
                           COL_VAT_RATE, COL_VAT_AMT, COL_SUM_VAT)

    last_data_row = DATA_START_ROW + len(items)
    _put_cell(ws, last_data_row, COL_N, "Всего")
    _put_qty_price_sum(ws, last_data_row, COL_QTY, COL_PRC, COL_SUM, total_qty, None, total_sum,
                       COL_VAT_RATE, COL_VAT_AMT, COL_SUM_VAT)

    # Рамки для верхнего блока (грузоотправитель, грузополучатель, поставщик, плательщик)
    _apply_border(ws, 3, 2, 15, 35)
    # Рамки для правого блока «Коды» (AK=37, AL=38, AM=39)
    _apply_border(ws, 2, 37, 17, 39)
    # Рамки для блока «Номер документа» / «Дата составления»
    _apply_border(ws, 16, 11, 17, 22)
    # Рамки для таблицы товаров (до колонки AM включительно)
    _apply_border(ws, 20, 2, last_data_row, 39)

    # Сумма прописью — «Всего отпущено на сумму» (строка 31)
    _put_cell(ws, 31, 2, f"Всего отпущено на сумму {_amount_to_words_rub(total_sum)}")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
